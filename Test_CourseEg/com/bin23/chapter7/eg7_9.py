import openpyxl

workbook = openpyxl.load_workbook('stock.xlsx')

sheetNames = workbook.sheetnames

for sheetName in sheetNames:
    print('当前工作表名称为：%s'%sheetName)
    # 根据sheetName获取sheet对象
    sheet = workbook[sheetName]
    # 遍历sheet每个单元格
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            print(sheet.cell(row = i + 1, column = j + 1).value, end = '\t')
        print()

# 也可以用下面的来遍历
# for i in range(sheet.max_row):
#     colName = 'A'
#     for j in range(sheet.max_column):
#         print(sheet['%s%d'%(colName, i + 1)].value, end = '\t')
#         colName = chr(ord(colName) + 1)
#     print()